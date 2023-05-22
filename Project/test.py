import openpyxl
from datetime import datetime
from salidas import drop_df, stock_df, dropEmergencias_df, stockEmergencias_df
from entradas import stock_entradas_df, stockEmergencias_entradas_df, drop_entradas_df, dropEmergencias_entradas_df
from resumenes import resumen_dropdf, resumen_entradasdf, resumen_stockdf

# Obtener la fecha actual en el formato MES AÑO
fecha_actual = datetime.now().strftime("%b %y")

# Crear un nuevo libro de Excel
nuevo_libro = openpyxl.Workbook()
nuevo_libro.remove(nuevo_libro["Sheet"])

# Crear la primera hoja con el nombre de fecha_actual
hoja_fecha = nuevo_libro.create_sheet(title=fecha_actual)

# Crear las otras hojas
hojas = ["ENTRADAS", "ENTRADAS EMERGENTES", "ENTRADAS DROP", "SALIDAS STOCK", "EMERGENCIAS STOCK", "EMERGENCIAS DROPS"]

for nombre_hoja in hojas:
    nueva_hoja = nuevo_libro.create_sheet(title=nombre_hoja)

# Guardar el archivo de Excel
ruta_nuevo_archivo = r"C:\Users\Frida Colin\MdkFacturacion\Archive\test.xlsx"
nuevo_libro.save(ruta_nuevo_archivo)


print("Archivo de Excel creado exitosamente.")
