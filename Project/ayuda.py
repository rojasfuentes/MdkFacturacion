from salidas import drop_df, stock_df, dropEmergencias_df, stockEmergencias_df
from entradas import stock_entradas_df, stockEmergencias_entradas_df, drop_entradas_df, dropEmergencias_entradas_df, numEle_stockEntradas , sumLineas_stockEntradas , sumUnidades_stockEntradas
from resumenes import resumen_dropdf, resumen_entradasdf, resumen_stockdf
from test import ruta_nuevo_archivo
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font


formato_path = ruta_nuevo_archivo

# Cargar el archivo de Excel existente con openpyxl
libro = load_workbook(formato_path)

# Crear un objeto ExcelWriter utilizando el archivo existente y el engine 'openpyxl'
writer = pd.ExcelWriter(formato_path, engine='openpyxl')
writer.book = libro

hoja_nombre = ''


# Obtener la hoja existente o crear una nueva si no existe
if 'ENTRADAS' in libro.sheetnames:
    hoja = libro['ENTRADAS']
else:
    hoja = libro.create_sheet(title=hoja_nombre)


# Convertir el dataframe a un objeto openpyxl para escribir en la hoja
data = stock_entradas_df.values.tolist()
# DATA con los nombres de las columnas
data.insert(0, list(stock_entradas_df.columns))
for row_index, row in enumerate(data):
    for col_index, value in enumerate(row):
        hoja.cell(row= 4 +row_index, column= 1 +col_index, value=value)
# Combinar las celdas A1 a N1
hoja.merge_cells('A1:N1')

# Escribir el texto en la celda A1
celda_a1 = hoja['A1']
celda_a1.value = 'MEDISTIK       * SH-SIEMENS HEALTHCARE DIAGS.*      LISTA DE CONTROL TIEMPO DE ENTRADAS'
celda_a1.alignment = Alignment(horizontal='center', vertical='center')
celda_a1.font = Font(bold=True)

conteo = hoja['A3']
conteo.value = numEle_stockEntradas
conteo.font = Font(bold=True)

sumLineas = hoja['G3']
sumLineas.value = sumLineas_stockEntradas
sumLineas.font = Font(bold=True)


sumUnidades = hoja['H3']
sumUnidades.value =  sumUnidades_stockEntradas
sumUnidades.font = Font(bold=True)
# Guardar los cambios en el archivo de Excel
libro.save(formato_path)

# Cerrar el objeto ExcelWriter
writer.close()